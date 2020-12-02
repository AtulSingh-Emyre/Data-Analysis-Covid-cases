from openpyxl import load_workbook
import math
import re
from numpy import *
from matplotlib import pyplot
from scipy.stats import *
wb = load_workbook('C:\\Users\\ashis\\Downloads\\owid-covid-data.xlsx')  # Work Book
ws = wb[wb.sheetnames[0]]
dataCol = "a b c d e f g h i j k l m n o p q r s t u v w x y z aa ab ac ad ae af ag ah ai aj ak al am an ao ap aq ar as at au av aw ax".upper().split(" ")
dataS5 = {}
for X in dataCol:
    column = ws[X]
    columnDate = ws['D']
    sample = []
    for x in range(1,len(columnDate)):
        test = columnDate[x].value
        test = test[len(columnDate[x].value)-2:]
        p = re.compile('01$')
        if p.match(test):
            sample.append(column[x].value)
    dataS5[column[0].value] = sample



columnHeadings = dataS5.keys()

def dailyCasesPlotter(code,key):
    iso_codes = dataS5['iso_code']
    date = dataS5['date']
    total_cases = dataS5[key]
    total_cases_IND = []
    date_IND = []
    for x in range(len(iso_codes)):
        if iso_codes[x] == code:
            total_cases_IND.append(total_cases[x])
            date_IND.append(date[x])
    pyplot.plot(date_IND,total_cases_IND)
    pyplot.title("Country : " + code)
    pyplot.xlabel("Days")
    pyplot.ylabel(key)
    pyplot.show()

countriesToLookFor = ['IND','USA']
for C in countriesToLookFor:    
    dailyCasesPlotter(C,'new_cases')
    


def dailyCasescomparison(code,key):
    iso_codes = dataS5['iso_code']
    date = dataS5['date']
    total_cases = dataS5[key]
    total_cases_IND = []
    date_IND = []
    for x in range(len(iso_codes)):
        if iso_codes[x] == code:
            total_cases_IND.append(total_cases[x])
            date_IND.append(date[x])
    return total_cases_IND

data=[]
countriesToLookFor = ['IND','USA']
for C in countriesToLookFor:    
    data.append(dailyCasescomparison(C,'new_cases'))
    
pyplot.boxplot(data)
pyplot.show()
