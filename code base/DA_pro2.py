from openpyxl import load_workbook
import math
from numpy import *
from matplotlib import pyplot 
from lmfit import Model
from scipy.stats import *
from scipy.optimize import curve_fit
wb = load_workbook("C:\\Users\\ashis\\Downloads\\DA data altitude.xlsx")  # Work Book
ws = wb[wb.sheetnames[0]]
column = ws['A']  # Column
columnalt = ws['D']
countriesToLookFor = ['India','Mexico','Nepal','United Kingdom','United States']
data=[]
countries=[]
for i in range(len(column)):
    countries.append(column[i].value)


for i in countriesToLookFor:
    x=columnalt[countries.index(i)].value
    s=''
    for i in x:
        if i.isdigit():
            s=s+i
        elif i==',':
            continue
        else:
            break
    data.append(int(s))

pyplot.bar(countriesToLookFor,data)
pyplot.title('altitude variations')
pyplot.ylabel('in meters')
pyplot.show()
