from openpyxl import load_workbook
import math
from numpy import *
from matplotlib import pyplot
from lmfit import Model
from scipy.stats import *
from scipy.optimize import curve_fit
wb = load_workbook("C:\\Users\\ashis\\Downloads\\Book1.xlsx")  # Work Book
ws = wb[wb.sheetnames[0]]
column = ws['A']
countries=['USA','IND']
time=['2019-Q1','2019-Q2','2019-Q3','2019-Q4','2020-Q1','2020-Q2']
values=ws['G']

gdpind=[]
gdpusa=[]
for j in range(1,len(column)):
    if column[j].value=='USA':
        gdpusa.append(values[j].value)
    if column[j].value=='IND':
        gdpind.append(values[j].value)


gdpusa.pop()
data=[gdpind,gdpusa]
fig, ax = pyplot.subplots() 
ax.bar(time, data[0], color = 'b', width = 0.25)
ax.bar(time, data[1], color = 'g', width = 0.25)
ax.legend(['IND','USA'],loc="lower left")
pyplot.xlabel('quarters')
pyplot.ylabel('GDP')
pyplot.show()
